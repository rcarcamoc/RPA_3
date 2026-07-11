#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sync_medicos_sharepoint.py
==========================
Sincroniza la tabla ris.medicos con los datos del archivo Excel en SharePoint.

Hoja fuente : "Validados por Integra"
Encabezados : fila 6
  Col B -> nombre_completo
  Col C -> usuario_integra
  Col D -> clave_integra

Logica de upsert:
  - Si usuario_integra NO existe -> INSERT
  - Si usuario_integra YA existe -> UPDATE si nombre o clave cambiaron

Estrategia de descarga (en orden de prioridad):
  1. MSAL ROPC (Azure AD con usuario/password - requiere que el tenant lo permita)
  2. SharePoint ADFS cookie-auth (tenants sin MFA)
  3. Archivo local (--local <ruta>)

Uso:
  python sync_medicos_sharepoint.py
  python sync_medicos_sharepoint.py --local "C:/ruta/al/archivo.xlsx"
"""

import sys
import os
import io
import re
import logging
import datetime
import argparse
import requests
import mysql.connector
import openpyxl

# ---- Forzar stdout UTF-8 en Windows ----------------------------------------
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ---- Logging ----------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("sync_medicos")

# ---- Credenciales SharePoint ------------------------------------------------
SP_TENANT     = "atryshealth0"                         # nombre del tenant
SP_TENANT_ID  = "atryshealth0.onmicrosoft.com"         # endpoint del tenant organizacion
SP_SITE_URL   = "https://atryshealth0-my.sharepoint.com"
SP_FILE_PATH  = "/personal/nsalinas_atryshealth_com/Documents/05. Ambulatorio Mayo 2026.xlsx"
SP_UNIQUE_ID  = "e64b460e-37f5-4383-8aad-84f69636ac0f"
SP_USER       = "cl-integramedica@atryshealth.com"
SP_PASS       = "4TRys.2024%"

# Client ID publico de SharePoint (app "Microsoft Office" - siempre disponible)
# Este es el client_id de la app de Microsoft que permite ROPC sin registrar app propia
SP_CLIENT_ID  = "d3590ed6-52b3-4102-aeff-aad2292ab01c"   # Microsoft Office

# Scopes para Graph API (para descargar archivos de OneDrive/SharePoint)
GRAPH_SCOPES  = ["https://graph.microsoft.com/.default"]
SP_SCOPES     = [f"{SP_SITE_URL}/.default"]

# ---- Configuracion Excel ----------------------------------------------------
NOMBRE_HOJA       = "Validados por Integra"
FILA_INICIO_DATOS = 7    # encabezados en fila 6, datos desde fila 7
COL_NOMBRE        = 2    # Columna B
COL_USUARIO       = 3    # Columna C
COL_CLAVE         = 4    # Columna D

# ---- Configuracion BD -------------------------------------------------------
DB_CONFIG = {
    "host":     "localhost",
    "user":     "root",
    "password": "",
    "database": "ris",
}


# =============================================================================
# DESCARGA DEL EXCEL DESDE SHAREPOINT
# =============================================================================

def _obtener_token_msal() -> str | None:
    """
    Obtiene access token via MSAL ROPC (Resource Owner Password Credentials).
    Retorna el token o None si falla.
    """
    try:
        import msal
    except ImportError:
        logger.warning("  msal no instalado. Ejecuta: pip install msal")
        return None

    authority = f"https://login.microsoftonline.com/{SP_TENANT_ID}"

    # Intentar primero con scopes de Graph API
    for scopes, nombre in [
        (["https://graph.microsoft.com/Files.Read.All"], "Graph API"),
        ([f"{SP_SITE_URL}/AllSites.Read"], "SharePoint"),
        (["https://graph.microsoft.com/.default"], "Graph default"),
    ]:
        logger.info(f"  Intentando token MSAL ({nombre})...")
        try:
            app = msal.PublicClientApplication(
                client_id=SP_CLIENT_ID,
                authority=authority,
            )
            result = app.acquire_token_by_username_password(
                username=SP_USER,
                password=SP_PASS,
                scopes=scopes,
            )

            if "access_token" in result:
                logger.info(f"  Token MSAL obtenido ({nombre})")
                return result["access_token"]
            else:
                error = result.get("error", "?")
                desc  = result.get("error_description", "")[:120]
                logger.warning(f"  MSAL fallo ({nombre}): {error} - {desc}")

        except Exception as e:
            logger.warning(f"  Error MSAL ({nombre}): {e}")

    return None


def _descargar_con_graph_api(token: str) -> bytes | None:
    """
    Descarga el archivo via Microsoft Graph API usando el token.
    Retorna bytes del xlsx o None.
    """
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/octet-stream",
    }

    # Construir URL de descarga de OneDrive/SharePoint via Graph API
    # Formato: /drives/{drive-id}/items/{item-id}/content
    # O via path: /users/{user}/drive/root:/{path}:/content
    user_email_encoded = SP_USER.replace("@", "%40")

    urls_graph = [
        # Por UniqueId (DriveItem)
        f"https://graph.microsoft.com/v1.0/drives/root/items/{SP_UNIQUE_ID}/content",
        # Por ruta del usuario
        f"https://graph.microsoft.com/v1.0/users/{SP_USER}/drive/root:"
        f"{SP_FILE_PATH}:/content",
        # Por SharedLink
        f"https://graph.microsoft.com/v1.0/shares/u!{_encode_sharing_url()}/driveItem/content",
    ]

    for url in urls_graph:
        logger.info(f"  Graph API: {url[:90]}...")
        try:
            resp = requests.get(url, headers=headers, timeout=60, allow_redirects=True)
            logger.info(f"  HTTP {resp.status_code}")
            if resp.status_code == 200 and resp.content[:4] == b'PK\x03\x04':
                logger.info(f"  Descarga exitosa via Graph API ({len(resp.content):,} bytes)")
                return resp.content
            elif resp.status_code == 200:
                logger.warning(f"  Respuesta no es Excel (bytes: {resp.content[:8].hex()})")
        except Exception as e:
            logger.warning(f"  Error: {e}")

    return None


def _encode_sharing_url() -> str:
    """Codifica la URL de SharePoint para Graph API sharing links."""
    import base64
    url = (
        "https://atryshealth0-my.sharepoint.com/:x:/r/personal/nsalinas_atryshealth_com"
        "/Documents/05.%20Ambulatorio%20Mayo%202026.xlsx"
        "?d=we64b460e37f543838aad84f69636ac0f&csf=1&web=1"
    )
    encoded = base64.urlsafe_b64encode(url.encode("utf-8")).decode("utf-8").rstrip("=")
    return encoded


def _descargar_con_sp_token(token: str) -> bytes | None:
    """
    Descarga el archivo via API REST de SharePoint usando el token.
    """
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/octet-stream",
    }

    urls = [
        (
            f"{SP_SITE_URL}/personal/nsalinas_atryshealth_com"
            f"/_layouts/15/download.aspx?UniqueId={SP_UNIQUE_ID}"
        ),
        (
            f"{SP_SITE_URL}/personal/nsalinas_atryshealth_com"
            f"/_api/web/GetFileByServerRelativePath(decodedurl='"
            f"/personal/nsalinas_atryshealth_com/Documents/05.%20Ambulatorio%20Mayo%202026.xlsx"
            f"')/$value"
        ),
    ]

    for url in urls:
        logger.info(f"  SP REST: {url[:90]}...")
        try:
            resp = requests.get(url, headers=headers, timeout=60, allow_redirects=True)
            logger.info(f"  HTTP {resp.status_code}")
            if resp.status_code == 200 and resp.content[:4] == b'PK\x03\x04':
                logger.info(f"  Descarga exitosa ({len(resp.content):,} bytes)")
                return resp.content
        except Exception as e:
            logger.warning(f"  Error: {e}")

    return None


def _descargar_con_adfs_cookies() -> bytes | None:
    """
    Intenta autenticacion ADFS (Forms Auth) de SharePoint Online.
    Funciona solo si el tenant no tiene MFA/Conditional Access bloqueando ROPC.
    """
    import xml.etree.ElementTree as ET

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 RPA-Sync/1.0"})

    sts_url = "https://login.microsoftonline.com/extSTS.srf"
    soap_body = f"""<?xml version="1.0" encoding="utf-8"?>
<s:Envelope xmlns:s="http://www.w3.org/2003/05/soap-envelope"
            xmlns:a="http://www.w3.org/2005/08/addressing"
            xmlns:u="http://docs.oasis-open.org/wss/2004/01/oasis-200401-wss-wssecurity-utility-1.0.xsd">
  <s:Header>
    <a:Action s:mustUnderstand="1">http://schemas.xmlsoap.org/ws/2005/02/trust/RST/Issue</a:Action>
    <a:ReplyTo><a:Address>http://www.w3.org/2005/08/addressing/anonymous</a:Address></a:ReplyTo>
    <a:To s:mustUnderstand="1">{sts_url}</a:To>
    <o:Security s:mustUnderstand="1"
      xmlns:o="http://docs.oasis-open.org/wss/2004/01/oasis-200401-wss-wssecurity-secext-1.0.xsd">
      <o:UsernameToken>
        <o:Username>{SP_USER}</o:Username>
        <o:Password>{SP_PASS}</o:Password>
      </o:UsernameToken>
    </o:Security>
  </s:Header>
  <s:Body>
    <t:RequestSecurityToken xmlns:t="http://schemas.xmlsoap.org/ws/2005/02/trust">
      <wsp:AppliesTo xmlns:wsp="http://schemas.xmlsoap.org/ws/2004/09/policy">
        <a:EndpointReference><a:Address>{SP_SITE_URL}/</a:Address></a:EndpointReference>
      </wsp:AppliesTo>
      <t:KeyType>http://schemas.xmlsoap.org/ws/2005/05/identity/NoProofKey</t:KeyType>
      <t:RequestType>http://schemas.xmlsoap.org/ws/2005/02/trust/Issue</t:RequestType>
      <t:TokenType>urn:oasis:names:tc:SAML:1.0:assertion</t:TokenType>
    </t:RequestSecurityToken>
  </s:Body>
</s:Envelope>"""

    try:
        logger.info("  ADFS: Obteniendo token STS...")
        r1 = session.post(sts_url, data=soap_body.encode("utf-8"),
                          headers={"Content-Type": "application/soap+xml; charset=utf-8"},
                          timeout=30)
        if r1.status_code != 200:
            return None

        root = ET.fromstring(r1.content)
        token_elem = root.find(
            ".//{http://docs.oasis-open.org/wss/2004/01/oasis-200401-wss-wssecurity-secext-1.0.xsd}"
            "BinarySecurityToken"
        )
        if token_elem is None or not token_elem.text:
            return None

        logger.info("  ADFS: Token STS OK. Autenticando en SharePoint...")
        signin_url = f"{SP_SITE_URL}/_forms/default.aspx?wa=wsignin1.0"
        session.post(signin_url, data=token_elem.text,
                     headers={"Content-Type": "application/x-www-form-urlencoded"},
                     timeout=30, allow_redirects=True)

        if "FedAuth" not in session.cookies and "rtFa" not in session.cookies:
            logger.warning("  ADFS: Cookies de sesion no obtenidas.")
            return None

        logger.info("  ADFS: Cookies OK. Descargando archivo...")
        dl_url = (
            f"{SP_SITE_URL}/personal/nsalinas_atryshealth_com"
            f"/_layouts/15/download.aspx?UniqueId={SP_UNIQUE_ID}"
        )
        r2 = session.get(dl_url, timeout=60, allow_redirects=True)
        if r2.status_code == 200 and r2.content[:4] == b'PK\x03\x04':
            logger.info(f"  ADFS: Descarga exitosa ({len(r2.content):,} bytes)")
            return r2.content

    except Exception as e:
        logger.warning(f"  ADFS error: {e}")

    return None


def descargar_excel_sharepoint() -> bytes:
    """
    Descarga el Excel probando multiples metodos de autenticacion.
    Lanza RuntimeError con instrucciones si todos fallan.
    """

    # Metodo 1: MSAL ROPC + Graph API
    logger.info("Metodo 1: MSAL (Azure AD ROPC) + Graph API...")
    token = _obtener_token_msal()
    if token:
        contenido = _descargar_con_graph_api(token)
        if contenido:
            return contenido
        # Tambien intentar con la API REST de SP con ese token
        contenido = _descargar_con_sp_token(token)
        if contenido:
            return contenido

    # Metodo 2: ADFS cookie-based
    logger.info("Metodo 2: ADFS cookie authentication...")
    contenido = _descargar_con_adfs_cookies()
    if contenido:
        return contenido

    # Todos los metodos fallaron
    raise RuntimeError(
        "\n"
        "No se pudo descargar el archivo de SharePoint automaticamente.\n"
        "\n"
        "SOLUCION RAPIDA - Usa el modo local:\n"
        "  1. Abre el link del Excel en tu navegador (ya tienes sesion activa)\n"
        "  2. Descarga el archivo manualmente (Archivo > Guardar una copia)\n"
        "  3. Ejecuta el script con el parametro --local:\n"
        "\n"
        "     python sync_medicos_sharepoint.py --local \"C:/ruta/archivo.xlsx\"\n"
        "\n"
        "SOLUCION PERMANENTE:\n"
        "  Si el tenant tiene MFA activo, se necesita registrar una App en Azure\n"
        "  AD y usar Client Credentials Flow. Consulta al administrador del tenant.\n"
    )


# =============================================================================
# LECTURA DEL EXCEL
# =============================================================================

def leer_medicos_excel(contenido_bytes: bytes) -> list:
    """
    Lee la hoja 'Validados por Integra' del Excel.
    Retorna lista de dicts con: nombre_completo, usuario_integra, clave_integra.
    Solo incluye filas donde usuario_integra no este vacio.
    """
    wb = openpyxl.load_workbook(io.BytesIO(contenido_bytes), data_only=True)

    if NOMBRE_HOJA not in wb.sheetnames:
        hojas = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Hoja '{NOMBRE_HOJA}' no encontrada en el Excel.\n"
            f"Hojas disponibles: {hojas}"
        )

    ws = wb[NOMBRE_HOJA]
    medicos = []
    total_vacias = 0

    logger.info(f"Hoja '{NOMBRE_HOJA}' - leyendo desde fila {FILA_INICIO_DATOS}...")

    for fila in ws.iter_rows(min_row=FILA_INICIO_DATOS, values_only=True):
        nombre  = fila[COL_NOMBRE  - 1] if len(fila) >= COL_NOMBRE  else None
        usuario = fila[COL_USUARIO - 1] if len(fila) >= COL_USUARIO else None
        clave   = fila[COL_CLAVE   - 1] if len(fila) >= COL_CLAVE   else None

        nombre  = str(nombre).strip()  if nombre  is not None else ""
        usuario = str(usuario).strip() if usuario is not None else ""
        clave   = str(clave).strip()   if clave   is not None else ""

        # Ignorar filas con usuario_integra vacio
        if not usuario or usuario.lower() in ("none", "nan"):
            total_vacias += 1
            continue

        medicos.append({
            "nombre_completo": nombre or usuario,
            "usuario_integra": usuario,
            "clave_integra":   clave,
        })

    logger.info(
        f"  {len(medicos)} registros validos | {total_vacias} filas vacias ignoradas"
    )
    return medicos


# =============================================================================
# NORMALIZACION
# =============================================================================

def _normalizar_nombre(nombre: str) -> str:
    """Minusculas, sin tildes, sin espacios dobles."""
    nombre = nombre.lower().strip()
    for orig, remp in [("\u00e1","a"),("\u00e9","e"),("\u00ed","i"),
                       ("\u00f3","o"),("\u00fa","u"),("\u00f1","n"),
                       ("\u00c1","a"),("\u00c9","e"),("\u00cd","i"),
                       ("\u00d3","o"),("\u00da","u"),("\u00d1","n")]:
        nombre = nombre.replace(orig, remp)
    return re.sub(r"\s+", " ", nombre).strip()


# =============================================================================
# UPSERT EN BASE DE DATOS
# =============================================================================

def sincronizar_con_bd(medicos: list) -> dict:
    """
    Hace upsert de los medicos en ris.medicos.
    Retorna estadisticas: insertados, actualizados, sin_cambios, errores.
    """
    stats = {"insertados": 0, "actualizados": 0, "sin_cambios": 0, "errores": 0}

    if not medicos:
        logger.warning("Lista de medicos vacia. Nada que sincronizar.")
        return stats

    try:
        conn   = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        logger.info(f"Conectado a MySQL: {DB_CONFIG['host']}/{DB_CONFIG['database']}")
    except mysql.connector.Error as e:
        logger.error(f"No se pudo conectar a la BD: {e}")
        raise

    try:
        for med in medicos:
            usuario  = med["usuario_integra"]
            nombre   = med["nombre_completo"]
            clave    = med["clave_integra"]
            nombre_n = _normalizar_nombre(nombre)

            try:
                cursor.execute(
                    "SELECT id_medico, nombre_completo, clave_integra "
                    "FROM medicos WHERE usuario_integra = %s LIMIT 1",
                    (usuario,),
                )
                existente = cursor.fetchone()

                if existente is None:
                    # INSERT
                    cursor.execute(
                        """
                        INSERT INTO medicos
                            (nombre_completo, usuario_integra, clave_integra)
                        VALUES (%s, %s, %s)
                        """,
                        (nombre, usuario, clave),
                    )
                    conn.commit()
                    logger.info(f"  [INSERT]  {usuario:35s} | {nombre}")
                    stats["insertados"] += 1

                else:
                    cambio_nombre = existente["nombre_completo"] != nombre
                    cambio_clave  = existente["clave_integra"]   != clave

                    if cambio_nombre or cambio_clave:
                        cursor.execute(
                            """
                            UPDATE medicos
                            SET nombre_completo = %s,
                                clave_integra   = %s
                            WHERE usuario_integra = %s
                            """,
                            (nombre, clave, usuario),
                        )
                        conn.commit()
                        cambios = []
                        if cambio_nombre:
                            cambios.append(
                                f"nombre '{existente['nombre_completo']}' -> '{nombre}'"
                            )
                        if cambio_clave:
                            cambios.append("clave [actualizada]")
                        logger.info(f"  [UPDATE]  {usuario:35s} | {', '.join(cambios)}")
                        stats["actualizados"] += 1

                    else:
                        logger.debug(f"  [OK]      {usuario}")
                        stats["sin_cambios"] += 1

            except mysql.connector.Error as e:
                logger.error(f"  [ERROR]   {usuario}: {e}")
                conn.rollback()
                stats["errores"] += 1

    finally:
        cursor.close()
        conn.close()
        logger.info("Conexion BD cerrada.")

    return stats


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Sincroniza medicos de SharePoint Excel -> ris.medicos",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Ejemplos:\n"
            "  python sync_medicos_sharepoint.py\n"
            "  python sync_medicos_sharepoint.py --local \"C:/Users/yo/Downloads/medicos.xlsx\"\n"
        ),
    )
    parser.add_argument(
        "--local",
        metavar="RUTA_EXCEL",
        help="Ruta al archivo Excel descargado manualmente (alternativa a la descarga automatica).",
    )
    args = parser.parse_args()

    inicio = datetime.datetime.now()
    sep = "=" * 65
    logger.info(sep)
    logger.info("SYNC MEDICOS SHAREPOINT -> ris.medicos")
    logger.info(f"Inicio: {inicio.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(sep)

    # ---- PASO 1: Obtener contenido del Excel --------------------------------
    if args.local:
        ruta = args.local
        if not os.path.isfile(ruta):
            logger.error(f"Archivo local no encontrado: {ruta}")
            sys.exit(1)
        logger.info(f"[PASO 1] Usando archivo local: {ruta}")
        with open(ruta, "rb") as f:
            contenido = f.read()
        logger.info(f"  Leido: {len(contenido):,} bytes")
    else:
        logger.info("[PASO 1] Descargando archivo Excel de SharePoint...")
        contenido = descargar_excel_sharepoint()

    # ---- PASO 2: Leer datos del Excel ---------------------------------------
    logger.info("[PASO 2] Leyendo datos del Excel...")
    medicos = leer_medicos_excel(contenido)

    if not medicos:
        logger.warning("No se encontraron medicos validos. Script finalizado.")
        sys.exit(0)

    # ---- PASO 3: Sincronizar con BD ----------------------------------------
    logger.info(f"[PASO 3] Sincronizando {len(medicos)} registros en ris.medicos...")
    stats = sincronizar_con_bd(medicos)

    # ---- Resumen ------------------------------------------------------------
    fin      = datetime.datetime.now()
    duracion = (fin - inicio).total_seconds()
    logger.info(sep)
    logger.info("RESUMEN")
    logger.info(f"  Insertados  : {stats['insertados']}")
    logger.info(f"  Actualizados: {stats['actualizados']}")
    logger.info(f"  Sin cambios : {stats['sin_cambios']}")
    logger.info(f"  Errores     : {stats['errores']}")
    logger.info(f"  Duracion    : {duracion:.1f}s")
    logger.info(sep)

    if stats["errores"] > 0:
        logger.warning(f"Finalizo con {stats['errores']} error(es).")
        sys.exit(1)
    else:
        logger.info("Sincronizacion completada exitosamente.")
        sys.exit(0)


if __name__ == "__main__":
    main()
