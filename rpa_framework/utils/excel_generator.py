import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
import mysql.connector
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Asegurar path
BASE_DIR = Path(__file__).resolve().parent.parent
REPORTS_DIR = BASE_DIR / "logs" / "reportes_excel"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def get_db_connection():
    try:
        from utils.mysql_auto_starter import ensure_mysql_running
        ensure_mysql_running()
    except Exception:
        pass
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="ris"
    )


def obtener_rango_fechas(periodo: str):
    """
    Retorna (fecha_inicio, fecha_fin, nombre_legible, sufijo_archivo)
    según el periodo solicitado: 'hoy', '7d', 'mes'
    """
    ahora = datetime.now()
    hoy_inicio = datetime.combine(ahora.date(), datetime.min.time())
    hoy_fin = datetime.combine(ahora.date(), datetime.max.time())

    if periodo == "hoy":
        fecha_inicio = hoy_inicio
        fecha_fin = hoy_fin
        nombre_legible = f"Día en curso ({ahora.strftime('%d/%m/%Y')})"
        sufijo_archivo = f"Hoy_{ahora.strftime('%Y%m%d')}"
    elif periodo == "7d":
        fecha_inicio = datetime.combine((ahora - timedelta(days=6)).date(), datetime.min.time())
        fecha_fin = hoy_fin
        nombre_legible = f"Últimos 7 días ({fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')})"
        sufijo_archivo = f"7Dias_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}"
    elif periodo == "mes":
        fecha_inicio = datetime(ahora.year, ahora.month, 1, 0, 0, 0)
        fecha_fin = hoy_fin
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        nombre_mes = meses[ahora.month - 1]
        nombre_legible = f"Mes actual ({nombre_mes} {ahora.year})"
        sufijo_archivo = f"Mes_{ahora.strftime('%Y%m')}"
    else:
        fecha_inicio = hoy_inicio
        fecha_fin = hoy_fin
        nombre_legible = f"Día en curso ({ahora.strftime('%d/%m/%Y')})"
        sufijo_archivo = f"Hoy_{ahora.strftime('%Y%m%d')}"

    return fecha_inicio, fecha_fin, nombre_legible, sufijo_archivo


def consultar_datos_periodo(fecha_inicio, fecha_fin):
    """Consulta la base de datos excluyendo tareas utilitarias internas."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
    SELECT 
        id,
        inicio,
        `update` AS fin,
        estado,
        numero_documento,
        examen,
        doctor_detectado,
        COALESCE(patologia_critica_detectada, patologia_critica) AS patologia_critica,
        fecha_actualizacion_notificacion,
        URL,
        diagnostico
    FROM registro_acciones
    WHERE inicio >= %s AND inicio <= %s
      AND (
          (ultimo_nodo IS NULL OR ultimo_nodo NOT IN ('Validación PACS', 'valida_pacs', 'seleccion int casos pendientes', 'casos_pendientes', 'Inicia RIS'))
          OR numero_documento IS NOT NULL
      )
      AND (observacion IS NULL OR (observacion NOT LIKE '%validación PACS%' AND observacion NOT LIKE '%validacion PACS%') OR numero_documento IS NOT NULL)
    ORDER BY id ASC
    """
    cursor.execute(query, (fecha_inicio, fecha_fin))
    filas = cursor.fetchall()
    cursor.close()
    conn.close()
    return filas


def generar_excel_reporte(periodo: str = "hoy"):
    """
    Genera un archivo Excel formateado profesionalmente para el periodo indicado.
    Retorna un diccionario con:
      - success (bool)
      - file_path (str)
      - filename (str)
      - nombre_periodo (str)
      - total_casos (int)
      - exitosos (int)
      - errores (int)
      - en_proceso (int)
      - patologias_criticas (int)
      - tasa_exito (str)
      - error_msg (str, opcional)
    """
    try:
        fecha_inicio, fecha_fin, nombre_periodo, sufijo = obtener_rango_fechas(periodo)
        datos = consultar_datos_periodo(fecha_inicio, fecha_fin)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Casos"
        ws.views.sheetView[0].showGridLines = True

        # Paleta de Estilos
        FONT_FAMILY = "Segoe UI"
        color_navy = "1F4E79"       # Azul corporativo
        color_header_fill = "1F4E79"
        color_kpi_bg = "F2F4F7"
        color_border = "D9D9D9"

        font_title = Font(name=FONT_FAMILY, size=16, bold=True, color="FFFFFF")
        font_subtitle = Font(name=FONT_FAMILY, size=10, italic=True, color="E0E0E0")
        font_kpi_label = Font(name=FONT_FAMILY, size=9, bold=False, color="595959")
        font_header = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        font_data = Font(name=FONT_FAMILY, size=9)
        font_link = Font(name=FONT_FAMILY, size=9, color="0563C1", underline="single")

        fill_title = PatternFill(start_color=color_navy, end_color=color_navy, fill_type="solid")
        fill_header = PatternFill(start_color=color_header_fill, end_color=color_header_fill, fill_type="solid")
        fill_kpi = PatternFill(start_color=color_kpi_bg, end_color=color_kpi_bg, fill_type="solid")
        fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Fills para Estados
        fill_exitoso = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        font_exitoso = Font(name=FONT_FAMILY, size=9, bold=True, color="276A3C")

        fill_error = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        font_error = Font(name=FONT_FAMILY, size=9, bold=True, color="C00000")

        fill_proceso = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        font_proceso = Font(name=FONT_FAMILY, size=9, bold=True, color="806000")

        fill_patologia = PatternFill(start_color="FFD8D8", end_color="FFD8D8", fill_type="solid")
        font_patologia = Font(name=FONT_FAMILY, size=9, bold=True, color="9C0006")

        thin_side = Side(border_style="thin", color=color_border)
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_kpi = Border(
            left=Side(border_style="thin", color="BDD7EE"),
            right=Side(border_style="thin", color="BDD7EE"),
            top=Side(border_style="thin", color="BDD7EE"),
            bottom=Side(border_style="thin", color="BDD7EE")
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # -------------------------------------------------------------
        # 1. Banner Principal (Filas 1 y 2)
        # -------------------------------------------------------------
        ws.merge_cells("A1:K1")
        ws.merge_cells("A2:K2")

        c1 = ws["A1"]
        c1.value = "  📊 REPORTE DE GESTIÓN RPA - ATRYS HEALTH"
        c1.font = font_title
        c1.fill = fill_title
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 32

        c2 = ws["A2"]
        c2.value = f"  Periodo: {nombre_periodo}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        c2.font = font_subtitle
        c2.fill = fill_title
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        for row in ws["A1:K2"]:
            for cell in row:
                cell.fill = fill_title

        # -------------------------------------------------------------
        # 2. Métricas Resumen (KPI Cards) (Filas 4 y 5)
        # -------------------------------------------------------------
        total_casos = len(datos)
        exitosos = 0
        errores = 0
        en_proceso = 0
        patologias = 0

        for r in datos:
            st = str(r.get("estado") or "").lower()
            if any(k in st for k in ['exito', 'exitoso', 'terminado', 'finalizado']):
                exitosos += 1
            elif any(k in st for k in ['error', 'fallo', 'falla']):
                errores += 1
            elif 'proceso' in st:
                en_proceso += 1
            else:
                if st: exitosos += 1

            if r.get("patologia_critica") and str(r.get("patologia_critica")).strip():
                patologias += 1

        tasa_exito = f"{(exitosos / total_casos * 100):.1f}%" if total_casos > 0 else "0.0%"

        kpis = [
            ("TOTAL CASOS", str(total_casos), "B4:C4", "B5:C5", "1F4E79"),
            ("EXITOSOS", f"{exitosos} ({tasa_exito})", "D4:E4", "D5:E5", "276A3C"),
            ("CON INCIDENCIAS", str(errores), "F4:G4", "F5:G5", "C00000"),
            ("EN PROCESO", str(en_proceso), "H4:I4", "H5:I5", "806000"),
            ("PATOLOGÍAS CRÍTICAS", str(patologias), "J4:K4", "J5:K5", "9C0006")
        ]

        ws.row_dimensions[4].height = 16
        ws.row_dimensions[5].height = 24

        for label, val, top_range, bot_range, col_hex in kpis:
            ws.merge_cells(top_range)
            ws.merge_cells(bot_range)

            top_cell = ws[top_range.split(":")[0]]
            top_cell.value = label
            top_cell.font = font_kpi_label
            top_cell.alignment = align_center

            bot_cell = ws[bot_range.split(":")[0]]
            bot_cell.value = val
            bot_cell.font = Font(name=FONT_FAMILY, size=12, bold=True, color=col_hex)
            bot_cell.alignment = align_center

            for cell_ref in [top_range, bot_range]:
                for r in ws[cell_ref]:
                    for c in r:
                        c.fill = fill_kpi
                        c.border = border_kpi

        # -------------------------------------------------------------
        # 3. Cabecera de Tabla (Fila 7)
        # -------------------------------------------------------------
        headers = [
            ("ID", 8, align_center),
            ("Inicio", 19, align_center),
            ("Fin / Update", 19, align_center),
            ("Estado", 14, align_center),
            ("N° Documento", 16, align_center),
            ("Examen", 32, align_left),
            ("Médico / Radiólogo", 28, align_left),
            ("Patología Crítica", 26, align_left),
            ("Fecha Gestión", 19, align_center),
            ("Enlace Informe", 35, align_left),
            ("Diagnóstico / Resumen", 45, align_wrap_left)
        ]

        HEADER_ROW = 7
        ws.row_dimensions[HEADER_ROW].height = 26

        for col_idx, (h_title, min_w, h_align) in enumerate(headers, 1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h_title)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all

        # -------------------------------------------------------------
        # 4. Filas de Datos
        # -------------------------------------------------------------
        current_row = HEADER_ROW + 1

        for idx, item in enumerate(datos):
            ws.row_dimensions[current_row].height = 22
            row_fill = fill_zebra if idx % 2 == 1 else fill_white

            f_ini = item.get("inicio")
            ini_str = f_ini.strftime("%d/%m/%Y %H:%M:%S") if hasattr(f_ini, "strftime") else (str(f_ini) if f_ini else "--")

            f_fin = item.get("fin")
            fin_str = f_fin.strftime("%d/%m/%Y %H:%M:%S") if hasattr(f_fin, "strftime") else (str(f_fin) if f_fin else "--")

            f_gest = item.get("fecha_actualizacion_notificacion")
            gest_str = f_gest.strftime("%d/%m/%Y %H:%M:%S") if hasattr(f_gest, "strftime") else (str(f_gest) if f_gest else "--")

            estado_val = str(item.get("estado") or "--")
            patologia_val = item.get("patologia_critica") or "--"
            url_val = item.get("URL") or "--"
            diag_val = item.get("diagnostico") or "--"

            if diag_val != "--":
                diag_val = "\n".join([line.strip() for line in str(diag_val).splitlines() if line.strip()])

            row_values = [
                (item.get("id"), align_center, None, None),
                (ini_str, align_center, None, None),
                (fin_str, align_center, None, None),
                (estado_val, align_center, None, None),
                (item.get("numero_documento") or "--", align_center, None, None),
                (item.get("examen") or "--", align_left, None, None),
                (item.get("doctor_detectado") or "--", align_left, None, None),
                (patologia_val, align_left, None, None),
                (gest_str, align_center, None, None),
                (url_val, align_left, None, None),
                (diag_val, align_wrap_left, None, None)
            ]

            for col_idx, (val, col_align, custom_font, custom_fill) in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = custom_font or font_data
                cell.fill = custom_fill or row_fill
                cell.alignment = col_align
                cell.border = border_all

                if col_idx == 4:  # Estado
                    st_low = estado_val.lower()
                    if any(k in st_low for k in ['exito', 'exitoso', 'terminado', 'finalizado']):
                        cell.fill = fill_exitoso
                        cell.font = font_exitoso
                    elif any(k in st_low for k in ['error', 'fallo', 'falla']):
                        cell.fill = fill_error
                        cell.font = font_error
                    elif 'proceso' in st_low:
                        cell.fill = fill_proceso
                        cell.font = font_proceso

                elif col_idx == 8:  # Patología Crítica
                    if patologia_val != "--" and patologia_val.strip():
                        cell.fill = fill_patologia
                        cell.font = font_patologia

                elif col_idx == 10:  # URL
                    if str(val).startswith("http"):
                        cell.font = font_link

            current_row += 1

        if total_casos == 0:
            ws.row_dimensions[current_row].height = 24
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            empty_cell = ws.cell(row=current_row, column=1, value="ℹ️ No se registraron casos en el periodo seleccionado.")
            empty_cell.font = Font(name=FONT_FAMILY, size=10, italic=True, color="7F7F7F")
            empty_cell.alignment = align_center
            empty_cell.fill = fill_white
            for c in ws[f"A{current_row}:K{current_row}"][0]:
                c.border = border_all
            current_row += 1

        # -------------------------------------------------------------
        # 5. Ajuste de Anchos y Filtro Automático
        # -------------------------------------------------------------
        for col_idx, (h_title, min_w, _) in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(min_w, len(h_title) + 4)

        last_data_row = max(current_row - 1, HEADER_ROW)
        ws.auto_filter.ref = f"A{HEADER_ROW}:K{last_data_row}"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Reporte_RPA_{sufijo}_{timestamp}.xlsx"
        filepath = REPORTS_DIR / filename
        wb.save(str(filepath))

        return {
            "success": True,
            "file_path": str(filepath),
            "filename": filename,
            "nombre_periodo": nombre_periodo,
            "total_casos": total_casos,
            "exitosos": exitosos,
            "errores": errores,
            "en_proceso": en_proceso,
            "patologias_criticas": patologias,
            "tasa_exito": tasa_exito
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error_msg": str(e)
        }


if __name__ == "__main__":
    print("Probando generador de Excel...")
    for p in ["hoy", "7d", "mes"]:
        res = generar_excel_reporte(p)
        print(f"Periodo '{p}': {res.get('success')} -> {res.get('file_path')}")
